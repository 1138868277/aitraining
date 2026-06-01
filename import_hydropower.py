"""导入水电编码字典到四个区域（使用 psql 命令行）"""
import openpyxl
import subprocess
import tempfile
import os
from datetime import datetime

EXCEL = "/Users/liuhaojun/Documents/AI培训/04 脚手架/aitraining/水电数据编码字典20260601(梳理后).xlsx"
NOW = datetime.now()

PG = ["psql", "-h", "localhost", "-U", "liuhaojun", "-v", "ON_ERROR_STOP=1", "-t", "-A"]

# 区域配置: (数据库, schema, 租户ID)
REGIONS = [
    ("training_exercises", "liuhaojun", "admin"),
    ("code_tools", "yunnan", "yunnan"),
    ("code_tools", "fujian", "fujian"),
    ("code_tools", "qianyuan", "qianyuan"),
]

def run_psql(dbname, sql):
    cmd = PG + ["-d", dbname, "-c", sql]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        print(f"  ERROR: {r.stderr}")
        raise Exception(f"psql error: {r.stderr}")
    return r.stdout.strip()

def main():
    wb = openpyxl.load_workbook(EXCEL)

    # === Sheet 2 ===
    ws2 = wb["二三级类对应关系"]
    second_set = {}
    third_list = []
    for r in range(2, ws2.max_row + 1):
        v1 = ws2.cell(r, 1).value
        v2 = ws2.cell(r, 2).value
        v3 = ws2.cell(r, 3).value
        v4 = ws2.cell(r, 4).value
        if not all([v1, v2, v3, v4]):
            continue
        sec_code = str(v1).strip().zfill(3)
        sec_name = str(v2).strip()
        third_code = str(v3).strip().zfill(3)
        third_name = str(v4).strip()
        second_set[(sec_code, sec_name)] = True
        third_list.append((sec_code, sec_name, third_code, third_name))

    second_items = sorted(second_set.items(), key=lambda x: x[0][0])

    # === Sheet 3 ===
    ws3 = wb["二级类-数据类-数据码"]
    code_list = []
    for r in range(2, ws3.max_row + 1):
        v1 = ws3.cell(r, 1).value
        v2 = ws3.cell(r, 2).value
        v3 = ws3.cell(r, 3).value
        v4 = ws3.cell(r, 4).value
        v5 = ws3.cell(r, 5).value
        v6 = ws3.cell(r, 6).value
        if not all([v1, v2, v3, v4, v5, v6]):
            continue
        sec_code = str(v1).strip().zfill(3)
        sec_name = str(v2).strip()
        cat_code = str(v3).strip().zfill(2)
        cat_name = str(v4).strip()
        data_code = str(v5).strip().zfill(3)
        data_name = str(v6).strip()
        code_list.append((sec_code, sec_name, cat_code, cat_name, data_code, data_name))

    print(f"Sheet2: {len(second_items)} 二级类, {len(third_list)} 三级类")
    print(f"Sheet3: {len(code_list)} 编码字典")

    for dbname, schema, tenant_id in REGIONS:
        print(f"\n===== 导入 {tenant_id} ({dbname}.{schema}) =====")

        # 获取最大ID
        max_type = int(run_psql(dbname, f"SELECT COALESCE(MAX(type_id),0) FROM {schema}.cec_new_energy_type_dict"))
        max_sec = int(run_psql(dbname, f"SELECT COALESCE(MAX(id),0) FROM {schema}.cec_new_energy_second_class_type_dict"))
        max_third = int(run_psql(dbname, f"SELECT COALESCE(MAX(third_class_id),0) FROM {schema}.cec_new_energy_third_class_dict"))
        max_code = int(run_psql(dbname, f"SELECT COALESCE(MAX(code_dict_id),0) FROM {schema}.cec_new_energy_code_dict"))
        print(f"  当前最大ID: type={max_type}, second={max_sec}, third={max_third}, code={max_code}")

        lines = ["BEGIN;"]
        now_str = NOW.strftime("%Y-%m-%d %H:%M:%S")

        # type_dict
        lines.append(f"INSERT INTO {schema}.cec_new_energy_type_dict (type_id,type_code,type_name,tenant_id,create_tm,modify_tm,if_delete) "
                     f"VALUES ({max_type+1},'S1','水力发电（常规）','{tenant_id}','{now_str}','{now_str}','0');")

        # second_class_type_dict
        sid = max_sec
        for (sc, sn), _ in second_items:
            sid += 1
            sn_esc = sn.replace("'", "''")
            lines.append(f"INSERT INTO {schema}.cec_new_energy_second_class_type_dict "
                         f"(id,type_code,second_class_code,second_class_name,tenant_id,create_tm,modify_tm,if_delete) "
                         f"VALUES ({sid},'S1','{sc}','{sn_esc}','{tenant_id}','{now_str}','{now_str}','0');")

        # third_class_dict
        tid = max_third
        for sc, sn, tc, tn in third_list:
            tid += 1
            sn_esc = sn.replace("'", "''")
            tn_esc = tn.replace("'", "''")
            lines.append(f"INSERT INTO {schema}.cec_new_energy_third_class_dict "
                         f"(third_class_id,type_code,second_class_code,second_class_name,third_class_code,third_class_name,tenant_id,create_tm,modify_tm,if_delete) "
                         f"VALUES ({tid},'S1','{sc}','{sn_esc}','{tc}','{tn_esc}','{tenant_id}','{now_str}','{now_str}','0');")

        # code_dict - batch insert
        cid = max_code
        for sc, sn, cc, cn, dc, dn in code_list:
            cid += 1
            sn_esc = sn.replace("'", "''")
            cn_esc = cn.replace("'", "''")
            dn_esc = dn.replace("'", "''")
            lines.append(f"INSERT INTO {schema}.cec_new_energy_code_dict "
                         f"(code_dict_id,type_domain_code,first_class_code,first_class_name,"
                         f"second_class_code,second_class_name,data_category_code,data_category_name,"
                         f"data_code,data_name,tenant_id,create_tm,modify_tm,if_delete,type_code) "
                         f"VALUES ({cid},'S','B1','生产运行','{sc}','{sn_esc}','{cc}','{cn_esc}',"
                         f"'{dc}','{dn_esc}','{tenant_id}','{now_str}','{now_str}','0','S1');")

        lines.append("COMMIT;")
        sql = "\n".join(lines)

        # 写入临时文件执行（避免shell转义问题）
        with tempfile.NamedTemporaryFile(mode="w", suffix=".sql", delete=False) as f:
            f.write(sql)
            tmp = f.name

        r = subprocess.run(PG + ["-d", dbname, "-f", tmp], capture_output=True, text=True)
        os.unlink(tmp)
        if r.returncode != 0:
            print(f"  ERROR: {r.stderr}")
            raise Exception(f"psql error: {r.stderr}")

        print(f"  type: 1, second: {len(second_items)}, third: {len(third_list)}, code: {len(code_list)}")
        print(f"  ✅ {tenant_id} 导入完成")

if __name__ == "__main__":
    main()
